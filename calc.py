from openpyxl import load_workbook
import tkinter as tk
wb = load_workbook('C:\\Users\\manojkiran\\Downloads\\roughdata.xlsx')
sheet = wb.active
def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 20

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"

def focus1(event):
    name_field.focus_set()

def focus2(event):
    course_field.focus_set()

def focus3(event):
    sem_field.focus_set()

def focus4(event):
    form_no_field.focus_set()

def focus5(event):
        contact_no_field.focus_set()

def clear():
    name_field.delete(0, tk.END)
    course_field.delete(0, tk.END)
    sem_field.delete(0, tk.END)
    form_no_field.delete(0, tk.END)
    contact_no_field.delete(0, tk.END)

def insert():

    if (name_field.get() == "" and
            course_field.get() == "" and
            sem_field.get() == "" and
            form_no_field.get() == "" and
            contact_no_field.get() == ""):

        print("empty input")

    else:

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        wb.save('C:\\Users\\manojkiran\\Downloads\\roughdata.xlsx')
        name_field.focus_set()
        clear()

if __name__ == "__main__":

    root = tk.Tk()
    root.configure(background='grey')
    root.title("registration form")
    root.geometry("500x300")
    excel()
    heading = tk.Label(root, text="Form", bg="grey")
    name = tk.Label(root, text="Name", bg="grey")
    course = tk.Label(root, text="Course", bg="grey")
    sem = tk.Label(root, text="Semester", bg="grey")
    form_no = tk.Label(root, text="Form No.", bg="grey")

    contact_no = tk.Label(root, text="Contact No.", bg="grey")
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)

    name_field = tk.Entry(root)
    course_field = tk.Entry(root)
    sem_field = tk.Entry(root)
    form_no_field = tk.Entry(root)
    contact_no_field = tk.Entry(root)
    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)

    sem_field.bind("<Return>", focus3)

    form_no_field.bind("<Return>", focus4)

    contact_no_field.bind("<Return>", focus5)

    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    excel()

    submit = tk.Button(root, text="Submit", fg="Black", bg="orange", command=insert)
    submit.grid(row=8, column=1)

    root.mainloop()